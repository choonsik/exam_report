import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
st.set_page_config(layout="wide", page_title="면접 심사 결과 리포트")

# --- Constants ---
# 점수 계산을 위한 카테고리별 컬럼 정의 (파일 양식에 맞게 수정)
# 이 컬럼들은 점수 계산 및 리포트 생성에 사용됩니다.
CATEGORY_COLS = {
    'Project': ['요구사항 관리', '사용방법론,도구', '목표달성/ 사업적 효과성'],
    'SW Architect': ['Architecting Process (접근방법 및 절차)', 'Architecture Design (표현 및 구조화)', 'Architecture 검증 (프로토타입 및 평가)'],
    'Communication': ['커뮤니케이션 (문서화/리더십)']
}

PASS_SCORE_THRESHOLD = 70
# 리포트 양식에 '제출용 양식' 추가
REPORT_FORMATS = ['상세 리포트', '요약 리포트', '제출용 양식']

# --- Excel Styling ---
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=12)
PASS_FONT = Font(color="0000FF", bold=True)
PASS_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
FAIL_FONT = Font(color="FF0000", bold=True)
FAIL_FILL = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
TABLE_HEADER_FONT = Font(bold=True)
TABLE_HEADER_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
BOX_BORDER = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))


# --- Helper Functions ---

def to_excel(df: pd.DataFrame) -> bytes:
    """데이터프레임을 엑셀 파일(bytes)로 변환합니다."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='통합결과')
    processed_data = output.getvalue()
    return processed_data

def apply_styles_to_range(ws, cell_range, font=None, fill=None, alignment=None, border=None):
    """주어진 범위의 셀에 스타일을 적용합니다. 단일 셀과 범위를 모두 처리합니다."""
    # 단일 셀 케이스를 명시적으로 처리
    if ':' not in cell_range:
        cell = ws[cell_range]
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        return

    # 범위 케이스 처리
    rows = ws[cell_range]
    # openpyxl이 단일 행 범위를 셀의 튜플로 반환하는 경우를 처리
    if rows and isinstance(rows[0], tuple):
        for row in rows:
            for cell in row:
                if font: cell.font = font
                if fill: cell.fill = fill
                if alignment: cell.alignment = alignment
                if border: cell.border = border
    # 단일 행 범위 처리
    else:
        for cell in rows:
            if font: cell.font = font
            if fill: cell.fill = fill
            if alignment: cell.alignment = alignment
            if border: cell.border = border

def write_individual_report_sheet(writer, candidate_name, all_df, report_format):
    """주어진 ExcelWriter 객체에 선택된 양식으로 개별 후보자의 리포트 시트를 작성하고 서식을 적용합니다."""
    sheet_name = f'{candidate_name} 리포트'
    
    # 데이터 준비
    candidate_df = all_df[all_df['성명'] == candidate_name].copy()
    is_final_pass = all(candidate_df['Reviewer_Result'] == 'Pass')
    final_result = "Pass" if is_final_pass else "Fail"
    
    # '제출용 양식' 로직
    if report_format == '제출용 양식':
        # 데이터프레임을 만들지 않고 빈 시트에 직접 작성
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 데이터 계산
        candidate_scores = candidate_df[list(CATEGORY_COLS.keys()) + ['총점']].mean()
        overall_avg = all_df[list(CATEGORY_COLS.keys()) + ['총점']].mean()
        passer_df = all_df[all_df['Reviewer_Result'] == 'Pass']
        passer_avg = passer_df[list(CATEGORY_COLS.keys()) + ['총점']].mean() if not passer_df.empty else pd.Series(0, index=candidate_scores.index)
        
        comments = candidate_df['총평'].fillna('코멘트 없음').tolist()

        # 셀에 데이터 쓰기
        worksheet['D1'] = "이름"
        worksheet['E1'] = candidate_name
        worksheet['B3'] = "Result"
        worksheet['C3'] = final_result
        
        worksheet['B5'] = "Category"
        worksheet['C5'] = "Total Average"
        worksheet['D5'] = "Average of pass applicants"
        worksheet['E5'] = candidate_name

        score_map = {
            'B6': 'Project (30)', 'C6': overall_avg['Project'], 'D6': passer_avg['Project'], 'E6': candidate_scores['Project'],
            'B7': 'SW Architect (60)', 'C7': overall_avg['SW Architect'], 'D7': passer_avg['SW Architect'], 'E7': candidate_scores['SW Architect'],
            'B8': 'Communication (10)', 'C8': overall_avg['Communication'], 'D8': passer_avg['Communication'], 'E8': candidate_scores['Communication'],
            'B9': 'Total (100)', 'C9': overall_avg['총점'], 'D9': passer_avg['총점'], 'E9': candidate_scores['총점'],
        }
        for cell, value in score_map.items():
            worksheet[cell] = value
            if isinstance(value, (int, float)):
                worksheet[cell].number_format = '0.00'

        for i in range(3):
            worksheet[f'B{11+i}'] = f'Reviewer{i+1} Comment'
            worksheet[f'C{11+i}'] = comments[i] if i < len(comments) else "N/A"

        # 서식 적용
        worksheet.column_dimensions['A'].width = 2
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 20
        
        # 셀 병합
        worksheet.merge_cells('C3:E3')
        for i in range(11, 14):
            worksheet.merge_cells(f'C{i}:E{i}')

        # 스타일 적용
        apply_styles_to_range(worksheet, 'D1', alignment=CENTER_ALIGN)
        apply_styles_to_range(worksheet, 'E1', alignment=CENTER_ALIGN)
        apply_styles_to_range(worksheet, 'B5:E9', border=THIN_BORDER)
        apply_styles_to_range(worksheet, 'B11:E13', border=THIN_BORDER)
        apply_styles_to_range(worksheet, 'B5:E5', font=TABLE_HEADER_FONT, alignment=CENTER_ALIGN)
        
        # C3 셀 서식 적용
        apply_styles_to_range(worksheet, 'C3', alignment=CENTER_ALIGN, border=THIN_BORDER)
        if final_result == "Pass":
            apply_styles_to_range(worksheet, 'C3', font=PASS_FONT, fill=PASS_FILL)
        else:
            apply_styles_to_range(worksheet, 'C3', font=FAIL_FONT, fill=FAIL_FILL)
        
        return # 제출용 양식은 여기서 종료

    # --- 기존 상세/요약 리포트 로직 ---
    comments_data = []
    candidate_df = candidate_df.reset_index(drop=True)
    for i, row in candidate_df.iterrows():
        reviewer_label = f"Reviewer {i+1}"
        result_label = f"(Pass)" if row['Reviewer_Result'] == 'Pass' else f"(Fail)"
        comment = row.get('총평', '코멘트 없음')
        comments_data.append({'심사위원': f"{reviewer_label} {result_label}", '코멘트': comment})
    comments_df = pd.DataFrame(comments_data)

    header_df = pd.DataFrame([{'항목': '후보자 리포트', ' ': candidate_name}, {'항목': '최종 결과', ' ': final_result}])
    header_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
    worksheet = writer.sheets[sheet_name]

    if report_format == '상세 리포트':
        candidate_scores = candidate_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("후보자 점수")
        overall_avg = all_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("전체 평균")
        passer_df = all_df[all_df['Reviewer_Result'] == 'Pass']
        passer_avg = passer_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("합격자 평균") if not passer_df.empty else pd.Series(0, index=candidate_scores.index, name="합격자 평균")
        comparison_df = pd.concat([candidate_scores, overall_avg, passer_avg], axis=1)
        comparison_df.index.name = "Category"
        
        worksheet['A4'] = '📊 심사 점수 분석'
        comparison_df.to_excel(writer, sheet_name=sheet_name, startrow=4)
        comments_start_row = 4 + len(comparison_df) + 3
    else: # 요약 리포트
        comments_start_row = 3

    worksheet[f'A{comments_start_row}'] = '📝 심사위원 코멘트'
    comments_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=comments_start_row)
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 80
    if report_format == '상세 리포트':
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15

    worksheet['A1'].font = TITLE_FONT
    worksheet['B1'].font = TITLE_FONT
    result_cell = worksheet['B2']
    if final_result == "Pass":
        result_cell.font = PASS_FONT
        result_cell.fill = PASS_FILL
    else:
        result_cell.font = FAIL_FONT
        result_cell.fill = FAIL_FILL
    result_cell.border = THIN_BORDER

    if report_format == '상세 리포트':
        worksheet['A4'].font = HEADER_FONT
        score_table_range = f'A5:D{5 + len(comparison_df)}'
        apply_styles_to_range(worksheet, score_table_range, border=THIN_BORDER)
        apply_styles_to_range(worksheet, f'A5:D5', font=TABLE_HEADER_FONT, fill=TABLE_HEADER_FILL, alignment=CENTER_ALIGN)

    worksheet[f'A{comments_start_row}'].font = HEADER_FONT
    comment_table_range = f'A{comments_start_row + 1}:B{comments_start_row + 1 + len(comments_df)}'
    apply_styles_to_range(worksheet, comment_table_range, border=THIN_BORDER)
    apply_styles_to_range(worksheet, f'A{comments_start_row + 1}:B{comments_start_row + 1}', font=TABLE_HEADER_FONT, fill=TABLE_HEADER_FILL, alignment=CENTER_ALIGN)


def generate_report_file_content(candidate_name, all_df, report_format):
    """선택된 후보자의 상세 리포트 내용을 Excel 파일(bytes)로 생성합니다."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        write_individual_report_sheet(writer, candidate_name, all_df, report_format)
    return output.getvalue()

def generate_overall_report_file_content(all_df, report_format):
    """전체 후보자에 대한 요약 및 개별 리포트를 포함하는 Excel 파일을 생성합니다."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. 전체 요약 시트 생성
        summary_data = []
        candidate_names = all_df['성명'].unique()
        for name in candidate_names:
            candidate_df = all_df[all_df['성명'] == name]
            is_final_pass = all(candidate_df['Reviewer_Result'] == 'Pass')
            final_result = "Pass" if is_final_pass else "Fail"
            avg_scores = candidate_df[list(CATEGORY_COLS.keys()) + ['총점']].mean()
            summary_row = {'성명': name, '최종 결과': final_result}
            summary_row.update(avg_scores)
            summary_data.append(summary_row)
        
        summary_df = pd.DataFrame(summary_data)
        # '전체 요약' 시트를 가장 먼저 생성
        summary_df.to_excel(writer, sheet_name='전체 요약', index=False)
        worksheet = writer.sheets['전체 요약']
        
        # 요약 시트 서식 적용
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        for col_idx, col_name in enumerate(summary_df.columns[2:], 3):
             worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
        
        summary_range = f'A1:{get_column_letter(len(summary_df.columns))}{len(summary_df) + 1}'
        apply_styles_to_range(worksheet, summary_range, border=THIN_BORDER)
        apply_styles_to_range(worksheet, f'A1:{get_column_letter(len(summary_df.columns))}1', font=TABLE_HEADER_FONT, fill=TABLE_HEADER_FILL, alignment=CENTER_ALIGN)
        
        # Pass/Fail 서식
        for row_idx, row in enumerate(summary_df.itertuples(), 2):
            result_cell = worksheet[f'B{row_idx}']
            if row._2 == "Pass":
                result_cell.font = PASS_FONT
                result_cell.fill = PASS_FILL
            else:
                result_cell.font = FAIL_FONT
                result_cell.fill = FAIL_FILL

        # 2. 후보자별 개별 리포트 시트 생성
        for name in candidate_names:
            write_individual_report_sheet(writer, name, all_df, report_format)
            
    return output.getvalue()


@st.cache_data
def load_and_process_data(uploaded_files):
    """
    업로드된 엑셀 파일들을 읽고 하나의 데이터프레임으로 통합 및 전처리합니다.
    - 입력 파일의 모든 컬럼을 유지합니다.
    - 각 파일의 '평가표' 시트를 읽습니다.
    - 5번째 행을 헤더로 사용하고, 데이터는 6번째 행부터 시작합니다.
    - 컬럼명의 개행 문자를 공백으로 변환합니다.
    - 누락된 '성명' 데이터를 제거합니다.
    - 카테고리별 점수와 Pass/Fail 여부를 계산하고, 원본 데이터와 비교 검증합니다.
    """
    if not uploaded_files:
        return pd.DataFrame()

    all_data = []
    for file in uploaded_files:
        try:
            # 첫 행을 1로 볼 때 5번째 행이 제목이므로, header 인덱스는 4가 됩니다.
            df = pd.read_excel(file, sheet_name='평가표', header=4)
            
            # 컬럼명에서 개행문자를 공백으로 변경하고 양쪽 공백을 제거합니다.
            df.columns = df.columns.str.replace('\n', ' ', regex=False).str.strip()
            
            # 모든 컬럼을 유지하므로, 별도의 컬럼 필터링을 하지 않습니다.
            all_data.append(df)
        except Exception as e:
            st.error(f"'{file.name}' 파일 처리 중 오류가 발생했습니다: {e}")
            st.info("엑셀 파일의 5번째 행에 컬럼명이 있고, '평가표' 시트가 존재하는지 확인해주세요.")
            return pd.DataFrame()

    if not all_data:
        return pd.DataFrame()

    # 모든 데이터프레임을 하나로 합칩니다.
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # '성명'이 비어있는 행은 제거합니다.
    combined_df.dropna(subset=['성명'], inplace=True)
    
    # --- 데이터 타입 변환 ---
    # 점수 계산에 필요한 컬럼들만 숫자 타입으로 변환합니다.
    score_cols = [col for sublist in CATEGORY_COLS.values() for col in sublist] + ['총점']
    
    for col in score_cols:
        # 파일에 해당 점수 컬럼이 있는 경우에만 변환 수행
        if col in combined_df.columns:
            combined_df[col] = pd.to_numeric(combined_df[col], errors='coerce').fillna(0)

    # --- 카테고리별 점수 및 Pass/Fail 계산 ---
    for category, cols in CATEGORY_COLS.items():
        # 파일에 존재하는 점수 컬럼만 합산
        score_cols_in_df = [c for c in cols if c in combined_df.columns]
        combined_df[category] = combined_df[score_cols_in_df].sum(axis=1)
    
    # '총점' 컬럼이 있는 경우에만 Pass/Fail 계산
    if '총점' in combined_df.columns:
        combined_df['Reviewer_Result'] = combined_df['총점'].apply(
            lambda x: 'Pass' if x >= PASS_SCORE_THRESHOLD else 'Fail'
        )
    else:
        # 총점 컬럼이 없으면 결과를 'N/A'로 처리
        combined_df['Reviewer_Result'] = 'N/A'
    
    # --- 합격여부 값 비교 검증 ---
    if '합격여부(Pass/Fail)' in combined_df.columns:
        # 비교를 위해 양쪽 값 정규화 (소문자, 공백 제거)
        original_result = combined_df['합격여부(Pass/Fail)'].astype(str).str.strip().str.lower()
        calculated_result = combined_df['Reviewer_Result'].str.strip().str.lower()
        
        # 원본 결과가 비어있지 않은 경우에만 비교하여 불일치 여부 플래그
        combined_df['Result_Mismatch'] = (original_result != calculated_result) & (original_result.notna()) & (original_result != '') & (original_result != 'nan')
    else:
        # 비교할 컬럼이 없으면 불일치 없음으로 처리
        combined_df['Result_Mismatch'] = False

    return combined_df

def generate_candidate_report(candidate_name, all_df):
    """선택된 후보자의 상세 리포트를 생성하고 다운로드 버튼을 제공합니다."""
    
    # 1. 후보자 데이터 추출
    candidate_df = all_df[all_df['성명'] == candidate_name].copy()
    if len(candidate_df) == 0:
        st.warning("해당 후보자의 데이터를 찾을 수 없습니다.")
        return

    # 2. 최종 합격 결과 계산
    is_final_pass = all(candidate_df['Reviewer_Result'] == 'Pass')
    final_result = "Pass" if is_final_pass else "Fail"
    
    st.header(f"👤 후보자 리포트: {candidate_name}")
    
    # 최종 결과와 다운로드 버튼을 나란히 배치
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        result_color = "blue" if final_result == "Pass" else "red"
        st.subheader(f"최종 결과: :{result_color}[{final_result}]")
    with col2:
        selected_format = st.radio("리포트 양식 선택", REPORT_FORMATS, horizontal=True, key=f"individual_report_format_{candidate_name}")
    with col3:
        st.write("") # 세로 정렬을 위한 빈 공간
        report_bytes = generate_report_file_content(candidate_name, all_df, selected_format)
        st.download_button(
            label="📥 리포트 다운로드",
            data=report_bytes,
            file_name=f"{candidate_name}_면접결과_{selected_format}.xlsx",
            mime="application/vnd.ms-excel"
        )

    st.markdown("---")

    # 3. 심사 점수 분석
    st.subheader("📊 심사 점수 분석")

    candidate_scores = candidate_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("후보자 점수")
    overall_avg = all_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("전체 평균")
    passer_df = all_df[all_df['Reviewer_Result'] == 'Pass']
    if not passer_df.empty:
        passer_avg = passer_df[list(CATEGORY_COLS.keys()) + ['총점']].mean().rename("합격자 평균")
    else:
        passer_avg = pd.Series(0, index=candidate_scores.index, name="합격자 평균")
    
    comparison_df = pd.concat([candidate_scores, overall_avg, passer_avg], axis=1)
    comparison_df.index.name = "Category"
    st.dataframe(comparison_df.style.format("{:.2f}"), use_container_width=True)
    
    st.markdown("---")

    # 4. 심사 리뷰 의견 (총평 사용)
    st.subheader("📝 심사위원 코멘트")
    candidate_df = candidate_df.reset_index(drop=True)
    for i, row in candidate_df.iterrows():
        reviewer_label = f"Reviewer {i+1}"
        result_label = f"({row['Reviewer_Result']})"
        
        with st.container(border=True):
            st.markdown(f"**{reviewer_label}** {result_label}")
            comment = row.get('총평', '코멘트 없음')
            st.info(f"{comment}")


# --- Streamlit App Main UI ---

st.title("📑 면접 심사 결과 리포트 생성기")
st.markdown("""
이 앱은 여러 개의 면접 심사표(Excel 파일)를 취합하여 후보자별 심사 결과를 분석하고 리포트를 생성합니다.
1.  **파일 업로드**: '평가표' 시트가 포함된 엑셀 파일들을 업로드하세요.
2.  **데이터 확인**: '통합 결과 확인' 탭에서 취합된 데이터를 확인하고 검증 문제를 확인하세요.
3.  **리포트 생성**: '후보자 리포트' 탭에서 특정 후보자를 선택하여 상세 리포트를 확인하고 다운로드하세요.
4.  **전체 리포트**: '전체 후보자 리포트' 탭에서 모든 후보자의 결과를 요약하고 전체 리포트를 다운로드하세요.
""")

uploaded_files = st.file_uploader(
    "면접 심사표 엑셀 파일을 업로드하세요.",
    type=['xlsx', 'xls'],
    accept_multiple_files=True
)

if uploaded_files:
    # 데이터 로드 및 처리
    processed_df = load_and_process_data(uploaded_files)

    if not processed_df.empty:
        # 탭 생성
        tab1, tab2, tab3 = st.tabs(["📊 통합 결과 확인", "📄 후보자 리포트", "🗂️ 전체 후보자 리포트"])

        with tab1:
            st.header("통합 심사 결과")

            # --- 데이터 검증 ---
            st.subheader("데이터 검증")
            evaluation_counts = processed_df.groupby('성명')['성명'].count()
            invalid_candidates = evaluation_counts[evaluation_counts != 3]

            if not invalid_candidates.empty:
                st.error("⚠️ **평가 횟수 오류**: 아래 후보자들은 3회의 평가를 받지 않았습니다.")
                st.dataframe(invalid_candidates.rename("평가 횟수"), use_container_width=True)
            else:
                st.success("✅ 모든 후보자가 3회의 평가를 받았습니다.")
            
            st.markdown("---")
            
            # --- 합격여부 결과 검증 ---
            st.subheader("합격/불합격 결과 검증")
            if 'Result_Mismatch' in processed_df.columns:
                mismatch_df = processed_df[processed_df['Result_Mismatch'] == True]
                
                if not mismatch_df.empty:
                    st.error("⚠️ **결과 불일치 오류**: 원본 파일의 합격 여부와 계산된 결과가 다릅니다.")
                    display_cols = ['성명', '심사위원 성명', '총점', '합격여부(Pass/Fail)', 'Reviewer_Result']
                    # 표시할 컬럼이 데이터프레임에 있는지 확인
                    display_cols = [col for col in display_cols if col in mismatch_df.columns]
                    st.dataframe(mismatch_df[display_cols], use_container_width=True)
                else:
                    st.success("✅ 모든 데이터의 합격 여부가 계산 결과와 일치합니다.")
            
            st.markdown("---")


            # --- 데이터 필터링 ---
            st.subheader("데이터 필터링 및 조회")
            
            # 필터링 UI
            col1, col2 = st.columns(2)
            with col1:
                candidate_names = sorted(processed_df['성명'].unique())
                selected_candidates = st.multiselect("후보자 선택", options=candidate_names, placeholder="모든 후보자 보기")
            with col2:
                # Reviewer_Result 컬럼이 존재하는 경우에만 필터 표시
                if 'Reviewer_Result' in processed_df.columns:
                    result_options = ['전체'] + processed_df['Reviewer_Result'].unique().tolist()
                    selected_result = st.selectbox("심사위원 평가 결과 선택", options=result_options, index=0)
                else:
                    selected_result = '전체'


            # 필터링 로직
            filtered_df = processed_df.copy()
            if selected_candidates:
                filtered_df = filtered_df[filtered_df['성명'].isin(selected_candidates)]
            if selected_result != '전체' and 'Reviewer_Result' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['Reviewer_Result'] == selected_result]

            st.dataframe(filtered_df.drop(columns=['Result_Mismatch'], errors='ignore'), use_container_width=True, hide_index=True)

            # --- 다운로드 버튼 ---
            st.download_button(
                label="📥 엑셀 파일로 다운로드",
                data=to_excel(filtered_df.drop(columns=['Result_Mismatch'], errors='ignore')),
                file_name="interview_results_combined.xlsx",
                mime="application/vnd.ms-excel"
            )

        with tab2:
            st.header("후보자별 상세 리포트")
            
            candidate_list = sorted(processed_df['성명'].unique())
            selected_candidate = st.selectbox(
                "리포트를 확인할 후보자를 선택하세요.",
                options=candidate_list,
                index=None,
                placeholder="후보자를 선택하세요"
            )

            if selected_candidate:
                generate_candidate_report(selected_candidate, processed_df)
        
        with tab3:
            st.header("전체 후보자 리포트 요약")

            summary_data = []
            candidate_names = sorted(processed_df['성명'].unique())
            for name in candidate_names:
                candidate_df = processed_df[processed_df['성명'] == name]
                is_final_pass = all(candidate_df['Reviewer_Result'] == 'Pass')
                final_result = "Pass" if is_final_pass else "Fail"
                avg_scores = candidate_df[list(CATEGORY_COLS.keys()) + ['총점']].mean()
                summary_row = {'성명': name, '최종 결과': final_result}
                summary_row.update(avg_scores)
                summary_data.append(summary_row)
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df.style.format("{:.2f}", subset=list(CATEGORY_COLS.keys()) + ['총점']), use_container_width=True, hide_index=True)
            
            col1, col2 = st.columns([1, 3])
            with col1:
                overall_report_format = st.radio("리포트 양식 선택", REPORT_FORMATS, horizontal=True, key="overall_report_format")
            with col2:
                st.download_button(
                    label="📥 전체 리포트 다운로드",
                    data=generate_overall_report_file_content(processed_df, overall_report_format),
                    file_name=f"interview_overall_{overall_report_format}.xlsx",
                    mime="application/vnd.ms-excel"
                )


else:
    st.info("심사 결과 분석을 시작하려면 엑셀 파일을 업로드해주세요.")
